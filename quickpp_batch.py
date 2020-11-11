import os
import shutil
import wave
import time
import glob
import re
import openpyxl
import gspread
from oauth2client.service_account import ServiceAccountCredentials 

sourceDirectory = 'c:\\quickpp'
excelChecklistName = 'data_collection_checklist_updates.xlsx'
excelChecklist = sourceDirectory + '\\' + excelChecklistName
scriptsDirectory = sourceDirectory + '\\Scripts'
mainSourceBranches = ['both', 'asr', 'wuw']

#CREDENTIALS
scope = ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
creds = ServiceAccountCredentials.from_json_keyfile_name(sourceDirectory+'json-credentials', scope)
try:
    client = gspread.authorize(creds)
except:
    print(' *** Cannot connect to Samsung User ID Mapping spreadsheet\nScripts will not be copied and files will not be checked')
    userPrompt = str(input('Continue (y/n)?'))
    if userPrompt != 'y':
        exit()

print('Quick Post-Process\n')
'''
while True:
    scriptNumber = int(input('Script number: '))
    if scriptNumber < 1 or scriptNumber > 75:
        print('Please enter valid script number (1-75)')
        continue
    else:
        break
while True:
    roomNumber = int(input('Room number: '))
    if roomNumber < 1 or roomNumber > 8:
        print('Please enter valid room number (1-8)')
        continue
    else:
        break
'''
startTime = time.time()

#fullScriptName = 'script' + str(scriptNumber).zfill(2) + '_condition' + str(roomNumber)
#print('Script name: ' + fullScriptName)

#make a list of raw sound folders (4 digit numbers)
rawDirectoryList = glob.glob('./[0-9][0-9][0-9][0-9]/')

for r in rawDirectoryList:
    os.chdir(r)

    rawDirectory = os.getcwd()
    rawDirectoryName = rawDirectory.split(os.sep)[-1]
    indexGrab = int(rawDirectoryName)+1
    
    try:
        #SHEET READING COMMAND 
        sheet = client.open('Test').sheet1
        # (INDEX,COLUMN) COLUMN WILL ALWAYS STAY AT 10 (SESSION/SCENARIO)
        fullScriptName = sheet.cell(indexGrab,10).value
        scriptInfo = re.findall(r'\d+', fullScriptName)
        roomNumber = scriptInfo[1]
        fullScriptName = 'script'+scriptInfo[0]+'_condition'+roomNumber
    except:
        print('\n *** Problem loading scripts from spreadsheet\nCheck internet connection and Samsung User ID Mapping spreadsheet')
    
    print('\nRaw sound directory detected: ' + rawDirectory)
    os.chdir('..')

    if os.path.isdir(rawDirectory + "_QA"):
        print('\nDirectory ' + rawDirectory + '_QA already exists. Check files and try again.')
        print(' *** Skipping conversion of ' + rawDirectory)
        continue

    os.mkdir(rawDirectory + '_QA')
    os.chdir(rawDirectory + '_QA')
    qaDirectory = os.getcwd()
    print('QA directory created: ' + qaDirectory)

    os.mkdir('recordings')
    
    try:
        scriptSrc = str(scriptsDirectory + '\\' + fullScriptName)
        scriptDst = str(qaDirectory + '\\' + fullScriptName)
        shutil.copytree(scriptSrc, scriptDst)
        print('Copying Scripts: ' + fullScriptName)
    except:
        print(' *** Script file not found at ' + scriptsDirectory + '\nContinuing without scripts...')
    

    def crawlDirectories(targetDir):
        print("Converting " + str(targetDir))
        directories = os.walk(targetDir)
        exclude = '510' # currently excluding files from Lux
        for root, dirs, files in directories:
            dirs[:] = [d for d in dirs if exclude not in d]

            getList = []
            for f in files:
                if ('wuw' in targetDir and "_mic_pcm" in f) or 'wuw' not in targetDir and (f.endswith(".pcm") or f.endswith(".raw")):
                    path = os.path.join(root, f)
                    getList.append(path)

            for list in getList:
                with open(list, 'rb') as pcmfile:
                    pcmdata = pcmfile.read()
                with wave.open(list+'.wav', 'wb') as wavfile:
                    wavfile.setparams((2, 2, 16000, 0, 'NONE', 'NONE'))
                    wavfile.writeframes(pcmdata)

                if 'wuw' in targetDir:
                    currentFile = list.split(os.sep)[-1]
                    destination = qaDirectory + "\\recordings\\" + findDirectoryWakeup(currentFile)
                else:
                    destination = qaDirectory + "\\recordings\\" + findDirectory(rawDirectory, root)
                
                if not os.path.isdir(destination):
                    os.mkdir(destination)
                shutil.copy(list+'.wav', destination)
                os.remove(list+'.wav')

    def findDirectory(source, target):
        # give the name of the set in raw to the destination qa directory
        sourceDir = source.split(os.sep)
        targetDir = target.split(os.sep)
        newTarget = targetDir[len(sourceDir) + 1]
        return newTarget

    def findDirectoryWakeup(source):
        # use the file name of the wake-up to set destination qa directory
        # [4] distance in meters
        # [6] None = clean
        # [10] Barge-in volume (non-bargeins have None for volume)
        splitName = source.split('_')
        if splitName[10] != 'None':
            return str(splitName[4] + ' barge-in ' + splitName[10])
        else:
            tvstatus = splitName[6] if splitName[6] == 'TV' else 'clean'
            #tvstatus = splitName[6] == 'TV' ? splitName[6] : 'clean'
            return str(splitName[4] + ' ' + tvstatus)
            
    def checkFiles():
        print('Counting files')
        os.chdir(rawDirectory + '_QA')
        checklist = openpyxl.load_workbook(excelChecklistName)
        sheet = checklist['room_' + str(roomNumber)]
        rows = [7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20] # ugh
        cols = ['B','C']
        
        for col in cols:
            for row in rows:
                r = str(row)
                placeholder = '<session_set>'
                if placeholder in sheet[col + r].value:
                    cellValue = str(sheet[col + r].value)
                    cellValue = cellValue.replace(placeholder, scriptInfo[0])
                    sheet[col + r].value = cellValue
        
        for row in rows:
            r = str(row)
            if sheet['C' + r].value == 'N/A':
                metersValue = sheet['E' + r].value
                typeValue = sheet['F' + r].value
                typeValue = typeValue.replace('(vol. ', '')
                typeValue = typeValue.replace(')', '')
                dir = os.getcwd() + '\\recordings\\' + metersValue + ' ' + typeValue
                
            else:
                dir = os.getcwd() + '\\recordings\\' + sheet['C' + r].value
                
            filecount = len([name for name in os.listdir(dir) if os.path.isfile(os.path.join(dir, name))])
            sheet['H' + r] = filecount
            if sheet['H' + r].value != sheet['G' + r].value:
                sheet['H' + r].fill = PatternFill(fgColor = 'FF0000', fill_type='solid')
                print(' *** Found file mismatch, check ' + dir)
            checklist.save(excelChecklistName)

    for branch in mainSourceBranches:
        crawlDirectories(rawDirectory + '\\' + branch)

    try:
        shutil.copy(excelChecklist, qaDirectory)
        print('Copying excel checklist')
        checkFiles()
    except:
        print(' *** Count failed, check files')
    
    os.chdir('..')

endTime = time.time()
print('\nCompleted ' + str(len(rawDirectoryList)) + ' operations in %s seconds' % round(endTime - startTime, 2))
input('Processing complete')
